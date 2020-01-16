import re, os
import xlrd
import numpy as np
import pandas as pd
from openpyxl import Workbook
from pyecharts import options as opts
from pyecharts.charts import Pie
from snapshot_selenium import snapshot as driver
from pyecharts.render import make_snapshot
from snapshot_selenium import snapshot

pd.set_option('display.width', None)
pd.set_option('display.max_columns', None)    # 设置显示所有列
pd.set_option('display.max_rows', None)

def average(data):
    return np.mean(data)

def bootstrap(data, B, C, func, cal_type='experience'):
    '''
    :param data: array数组保存数据
    :param B: 抽样次数，通常 B >= 1000
    :param C: 置信水平
    :param func: 样本估计量
    :param cal_type: bootstrap 计算类型，经验法和百分位法，默认使用经验法
    :return: bootstrap 置信区间上下限
    '''
    array = np.array(data)
    aver = func(data)
    # print(aver)
    n = len(array)
    result = []
    for _ in range(B):
        index_arr = np.random.randint(0, n, size=n)
        sample = array[index_arr]
        sample_result = func(sample)
        result.append(sample_result)

    if cal_type == 'experience':
        a = 1 - C
        k1 = int(B * a / 2)
        k2 = int(B * (1 - a / 2))
        result_sorted = sorted(result)
        lower = round(aver - result_sorted[k1], 3)
        upper = round(aver + result_sorted[k2], 3)
        return [lower, upper]
    else:
        # 百分位法
        lower_p = (100 - C * 100) / 2
        # lower = round(max(0.0, np.percentile(result, lower_p)), 3)
        lower = round(np.percentile(result, lower_p), 3)
        upper_p = 100 * C + lower_p
        # upper = round(min(1.0, np.percentile(result, upper_p)), 3)
        upper = round(np.percentile(result, upper_p), 3)
        return [lower, upper]

def process_data(file):
    df = pd.read_csv(file, sep='\t', index_col=0)
    newdf = df.T
    normal_list = [i for i in df.columns if re.search('B', i)]
    df_normal = df[normal_list].T

    return newdf, df_normal

def sample_info(file):
    with open(file, encoding='utf8') as f:
        sample_info_dict = {line.strip().split('\t')[0]: line.strip().split('\t')[1] for line in f}
    return sample_info_dict

def write(data1, data2, sample_name, ci1, ci2):
    '''
    :param data:  dataframe数据
    :param sample: 样本编号
    # :param sample_dict: 样本编号对应的病人信息
    :param confidence_interval: 置信区间
    :return:
    '''
    # model = xlrd.open_workbook('C:\\Users\\jbwang\\Desktop\\P20190700965-SX1\\model.xlsx').sheets()[0]

    helpful_list = ['栖粪杆菌属', '双歧杆菌属', '拟杆菌属', '瘤胃球菌属-2', '罗斯氏菌属', '丁酸弧菌', '艾克曼菌属']
    harmful_list = ['大肠-志贺氏菌属', '链球菌属', '柯林斯氏菌属']

    data1 = pd.DataFrame(data1)
    data1.columns = ['您']
    data1['健康人(百分比)'] = [f'{i[0]}-{i[1]}' for i in list(ci1.values())]

    # data_help = data1.loc[helpful_list, :]
    # data_harm = data1.loc[harmful_list, :]

    ## 多样性
    data2 = pd.DataFrame(data2)
    data2.columns = ['您']
    data2['健康人(百分比)'] = [f'{i[0]}-{i[1]}' for i in list(ci2.values())]

    wb = Workbook()
    # ws = wb.active
    ws1 = wb.create_sheet('Top20', 0)
    ws2 = wb.create_sheet('益生菌', 1)
    ws3 = wb.create_sheet('有害菌', 2)
    ws4 = wb.create_sheet('多样性', 3)

    title = ['名称', '您', '健康人(百分比)']
    ws1.append(title)
    ws2.append(title)
    ws3.append(title)
    ws4.append(['名称', '您', '健康人'])

    def myfunc(data, i):
        you = data.loc[i, '您']
        normal = data.loc[i, '健康人(百分比)']
        if float(you) < float(normal.split('-')[0]):
            you = f'{you}↓'
        elif float(you) > float(normal.split('-')[1]):
            you = f'{you}↑'

        tmp_line = f'{i},{you},{normal}'.split(',')
        return tmp_line

    for i in data1.index:
        line1 = myfunc(data1, i)
        ws1.append(line1)
        if i in helpful_list:
            ws2.append(line1)
        elif i in harmful_list:
            ws3.append(line1)

    for j in data2.index:
        line2 = myfunc(data2, j)
        ws4.append(line2)

    wb.save(f'C:\\Users\\jbwang\\Desktop\\P20190700965-SX2\\报告及附件\\{sample_name}\\result.xlsx')

def pie(data, sample_name):
    lable = data.index.tolist()
    def pie_radius() -> Pie:
        c = (
            Pie(init_opts=opts.InitOpts(width="1000px", height="600px"))
                .add(
                "",
                [list(z) for z in zip(lable, data)],
                radius=["40%", "75%"],
                #center=["25%", "50%"]
            )
                .set_global_opts(
                title_opts=opts.TitleOpts(title="您", pos_left='47.5%', pos_top='46%',
                                          title_textstyle_opts=opts.series_options.TextStyleOpts(font_size=40)),
                legend_opts=opts.LegendOpts(is_show=False)
            )
            #         .set_series_opts(label_opts=opts.LabelOpts(formatter="{b}"))
        )
        return c

    pie_radius().render(f'C:\\Users\\jbwang\\Desktop\\P20190700965-SX2\\报告及附件\\{sample_name}\\pie.html')
    # make_snapshot(snapshot, pie_radius().render(), f'C:\\Users\\jbwang\\Desktop\\P20190700965-SX2\\报告及附件\\{sample_name}\\pie.png')
    make_snapshot(snapshot, pie_radius().render(), f'C:\\Users\\jbwang\\Desktop\\P20190700965-SX2\\报告及附件\\{sample_name}\\pie.pdf')

def main():
    data = process_data('C:\\Users\\jbwang\\Desktop\\P20190700965-SX2\\top20.txt')
    data_index = process_data('C:\\Users\\jbwang\\Desktop\\P20190700965-SX2\\index.txt')
    # print(data_index[1])
    sample_dict = sample_info('C:\\Users\\jbwang\\Desktop\\P20190700965-SX2\\sample_info.txt')
    confidence_interval = {i: bootstrap(data[1][i], 1000, 0.90, average, cal_type='percential') for i in data[1].columns}
    ci_index = {i: bootstrap(data_index[1][i], 1000, 0.90, average, cal_type='percential') for i in data_index[1].columns}

    print(confidence_interval)
    print(ci_index)
    for s in data[0].index:
        tmp_data1 = data[0].loc[s, :]
        tmp_data2 = data_index[0].loc[s, :]
        # print(tmp_data2)
        sample_name = sample_dict.get(s)
        if os.path.exists(f'C:\\Users\\jbwang\\Desktop\\P20190700965-SX2\\报告及附件\\{sample_name}'):
            # pie(tmp_data1, sample_name)
            write(tmp_data1, tmp_data2, sample_name, confidence_interval, ci_index)
        else:
            os.mkdir(f'C:\\Users\\jbwang\\Desktop\\P20190700965-SX2\\报告及附件\\{sample_name}')
            # pie(tmp_data1, sample_name)
            write(tmp_data1, tmp_data2, sample_name, confidence_interval, ci_index)


if __name__ == '__main__':
    main()


