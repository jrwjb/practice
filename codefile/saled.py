import re
import json
import xlrd
import numpy as np
import pandas as pd
from openpyxl import Workbook

pd.set_option('display.width', None)
pd.set_option('display.max_columns', None)    # 设置显示所有列
pd.set_option('display.max_rows', None)      # 显示所有行

############## R20190300352-SX1 ###############
# def gene_name():
#     gene_dict = {}
#     with open('E:\\Project\\Saled\\R20190300352-SX1\\1.txt') as f:
#         for i in f:
#             protein = i.split('|')[1]
#             if re.search('GN=\S+', i):
#                 gene = re.search('GN=\S+', i).group().split('=')[1]
#             else:
#                 gene = ''
#             gene_dict[protein] = gene
#             # f2.write(f'{protein},{gene}\n')
#     return gene_dict
#
#
# go = xlrd.open_workbook('E:\\Project\\Saled\\R20190300352-SX1\\R20190300352_RBP1_4.xlsx')
# go_table = go.sheet_by_index(0)
#
# wb = Workbook()
# ws = wb.active
# ws.append(go_table.row_values(0) + ['GeneName'])
# ws.append(go_table.row_values(1))
#
# # for row in range(1, go_table.nrows):
# #     line = go_table.row_values(row)
# #     proteinID = line[0]
# #     # print(proteinID)
# #     if proteinID in gene_name():
# #         geneName = gene_name()[proteinID]
# #         line.append(geneName)
# #     ws.append(line)
#
# for row in range(2, go_table.nrows):
#     line = go_table.row_values(row)
#     if str(line[0]).endswith('-1'):
#         proteinID = line[1].split('|')[1]
#     # print(proteinID)
#         if proteinID in gene_name():
#             line.append(gene_name()[proteinID])
#         ws.append(line)
#     else:
#         ws.append(line)
#
#
# wb.save('E:\\Project\\Saled\\R20190300352-SX1\\GO2.xlsx')

# gene_id = pd.read_csv('E:\\Project\\Saled\\R20190300352-SX1\\2.csv', header=None, index_col=0)
# print(gene_id.head())

# data = pd.read_excel('E:\\Project\\Saled\\R20190300352-SX1\\R20190300352_RBP1_2.xlsx', sheet_name=0)
# index = [i for i in data.index if str(i).endswith('-1')]
# df = data.loc[index]
# # df['tmp'] = df.index
# df = df.reset_index(drop=True)
# new_index = [i.split('|')[1] for i in df['Reference']]
# df.index = new_index
#
# new_df = gene_id.loc[new_index]
# new_df.to_csv('E:\\Project\\Saled\\R20190300352-SX1\\1.csv')
# # # print(df.head())

############### P20181000755-SX1 #####################

# group = 'CK60d_vs_CK33d'
# #
# ### 更改proteinID 为 protein name
# protein = pd.read_csv('E:\\Project\\Saled\\P20181000755-SX1\\protein.txt', sep='\t')
# protein_dict = {k: v for k, v in zip(protein['Protein'], protein['Protein Name'])}
# # print(protein_dict)
#
# up_file = pd.read_csv(f'E:\\Project\\Saled\\P20181000755-SX1\\{group}\\up.txt', sep='\t', header=None)
# down_file = pd.read_csv(f'E:\\Project\\Saled\\P20181000755-SX1\\{group}\\down.txt', sep='\t', header=None)
# ko_file = pd.read_csv(f'E:\\Project\\Saled\\P20181000755-SX1\\{group}\\query.ko', sep='\t', header=None)
# columns = ['Pathway_ID', 'Pathway', 'level1', 'level2', 'Up_number', 'Down_number', 'DEG_number',
#                'Total_number', 'pvalue', 'FDR', 'Up_gene', 'Down_gene', 'Up_KO', 'Down_KO', 'URL']
#
# ko_dict = {k: v for k, v in zip(ko_file[0], ko_file[1])}
# # print(ko_dict)
#
# def count_num(up, down, seqs):
#     up_num, down_num, up_gene_list, down_gene_list = [], [], [], []
#     for s in seqs:
#         c_up = c_down = 0
#         up_gene, down_gene = [], []
#         for u, d in zip(up, down):
#             if u in s:
#                 c_up += 1
#                 # up_gene.append(u)
#                 up_gene.append(protein_dict[u])   ### 更改proteinID为protein name
#             elif d in s:
#                 c_down += 1
#                 # down_gene.append(d)
#                 down_gene.append(protein_dict[d])
#         up_num.append(c_up)
#         down_num.append(c_down)
#         up_gene_list.append(';'.join(up_gene))
#         down_gene_list.append(';'.join(down_gene))
#     return up_num, down_num, up_gene_list, down_gene_list
#
# ##### GO
# def go_anno(index):
#     go = pd.read_excel(f'E:\\Project\\Saled\\P20181000755-SX1\\{group}\\go\\GO.xlsx', sheet_name=index)
#     df = pd.DataFrame(columns=columns)
#     df['Pathway_ID'] = go['GO_ID']
#     df['Pathway'] = go['Term']
#     if index != 3:
#         go_seqs = go['Sequences\n']
#         up_number = count_num(up_file[0], down_file[0], go_seqs)[0]
#         down_number = count_num(up_file[0], down_file[0], go_seqs)[1]
#         deg_number = np.array(up_number) + np.array(down_number)
#         up_data = count_num(up_file[0], down_file[0], go_seqs)[2]
#         down_data = count_num(up_file[0], down_file[0], go_seqs)[3]
#         df['Up_number'] = up_number
#         df['Down_number'] = down_number
#         df['DEG_number'] = deg_number
#         df['Up_gene'] = up_data
#         df['Down_gene'] = down_data
#     else:
#         go_seqs = go['TestSeqs']
#         up_number = count_num(up_file[0], down_file[0], go_seqs)[0]
#         down_number = count_num(up_file[0], down_file[0], go_seqs)[1]
#         deg_number = np.array(up_number) + np.array(down_number)
#         up_data = count_num(up_file[0], down_file[0], go_seqs)[2]
#         down_data = count_num(up_file[0], down_file[0], go_seqs)[3]
#         df['Up_number'] = up_number
#         df['Down_number'] = down_number
#         df['DEG_number'] = deg_number
#         df['Total_number'] = go['Ref']
#         df['pvalue'] = go['P value']
#         df['FDR'] = go['FDR']
#         df['Up_gene'] = up_data
#         df['Down_gene'] = down_data
#
#     return df
#
# newdf = pd.concat([go_anno(0), go_anno(1), go_anno(2)])
# newdf2 = go_anno(3)
# # pandas写入俩个sheet
# writer = pd.ExcelWriter(f'E:\\Project\\Saled\\P20181000755-SX1\\报告及附件\\{group}\\go\\GO.xlsx')
# newdf.to_excel(writer, index=False, sheet_name='Annotation')
# newdf2.to_excel(writer, index=False, sheet_name='Enrichment')
# writer.save()
# writer.close()
# # print(newdf.head())
#
# ###### kegg
#
# def kegg_anno(index):
#     kegg = pd.read_excel(f'E:\\Project\\Saled\\P20181000755-SX1\\{group}\\kegg\\kegg.xlsx', sheet_name=index)
#     kegg_df = pd.DataFrame(columns=columns)
#     kegg_df['Pathway_ID'] = kegg['Map_ID']
#     kegg_df['Pathway'] = kegg['Map_Name']
#
#     def ko_func(x):
#         ko_list = []
#         for p in x:
#             p_id_list = p.split(';')
#             ko = []
#             for p_id in p_id_list:
#                 if p_id in ko_dict:
#                     ko.append(ko_dict[p_id])
#             ko_list.append(';'.join(ko))
#         return ko_list
#
#     def level():
#         with open('C:\\Users\\jbwang\\Desktop\\br08901.json', 'r') as f:
#             data = json.load(f)
#             kegg_map = data['children']
#             level1_list, level2_list = [], []
#             for kegg_id in kegg['Map_ID']:
#                 for i in kegg_map:
#                     A = i['name']
#                     B_list = i['children']
#                     for j in B_list:
#                         B = j['name']
#                         C_list = j['children']
#                         for k in C_list:
#                             C = k['name'].split()[0]
#                             if C == kegg_id[3:]:
#                                 level1_list.append(A)
#                                 level2_list.append(B)
#         return level1_list, level2_list
#
#     if index == 1:
#         kegg_seqs = kegg['Seqs']
#         k_up_number = count_num(up_file[0], down_file[0], kegg_seqs)[0]
#         k_down_number = count_num(up_file[0], down_file[0], kegg_seqs)[1]
#         k_deg_number = np.array(k_up_number) + np.array(k_down_number)
#         k_up_data = count_num(up_file[0], down_file[0], kegg_seqs)[2]
#         k_down_data = count_num(up_file[0], down_file[0], kegg_seqs)[3]
#         kegg_df['Up_KO'] = ko_func(k_up_data)
#         kegg_df['Down_KO'] = ko_func(k_down_data)
#         kegg_df['Up_number'] = k_up_number
#         kegg_df['Down_number'] = k_down_number
#         kegg_df['DEG_number'] = k_deg_number
#         kegg_df['Up_gene'] = k_up_data
#         kegg_df['Down_gene'] = k_down_data
#         kegg_df['level1'] = level()[0]
#         kegg_df['level2'] = level()[1]
#         kegg_df['URL'] = kegg['URL']
#     else:
#         kegg_seqs = kegg['Test_Seq']
#         k_up_number = count_num(up_file[0], down_file[0], kegg_seqs)[0]
#         k_down_number = count_num(up_file[0], down_file[0], kegg_seqs)[1]
#         k_deg_number = np.array(k_up_number) + np.array(k_down_number)
#         k_up_data = count_num(up_file[0], down_file[0], kegg_seqs)[2]
#         k_down_data = count_num(up_file[0], down_file[0], kegg_seqs)[3]
#         kegg_df['Up_KO'] = ko_func(k_up_data)
#         kegg_df['Down_KO'] = ko_func(k_down_data)
#         kegg_df['Up_number'] = k_up_number
#         kegg_df['Down_number'] = k_down_number
#         kegg_df['DEG_number'] = k_deg_number
#         kegg_df['Up_gene'] = k_up_data
#         kegg_df['Down_gene'] = k_down_data
#         kegg_df['level1'] = level()[0]
#         kegg_df['level2'] = level()[1]
#         kegg_df['Total_number'] = kegg['Ref']
#         kegg_df['pvalue'] = kegg['P value']
#         kegg_df['FDR'] = kegg['FDR']
#     return kegg_df
#
# # print(kegg_df.head())
# writer = pd.ExcelWriter(f'E:\\Project\\Saled\\P20181000755-SX1\\报告及附件\\{group}\\kegg\\kegg.xlsx')
# kegg_df1 = kegg_anno(1)
# kegg_df2 = kegg_anno(2)
# kegg_df1.to_excel(writer, index=False, sheet_name='Annotation')
# kegg_df2.to_excel(writer, index=False, sheet_name='Enrichment')
#
# writer.save()
# writer.close()

#### P20190100016-SX1

# data = xlrd.open_workbook('E:\\Project\\Saled\\P20190100016-SX1\\附件1_磷酸化肽段鉴定列表.xlsx')
# data_table = data.sheet_by_index(0)
#
# wb = Workbook()
# ws = wb.active
# gene_dict = {}
# for row in range(1, data_table.nrows):
#     line = data_table.row_values(row)
#     seqs = line[0]
#     proteinID = line[2].split(';')[0]
#     mod = line[6].replace(' ', '')
#     tmp = proteinID + '|' + mod + '|' + seqs
#     gene = line[4] + '|' + mod + '|' + seqs
#     gene_dict[tmp] = gene
#
# ### cluster
# f1 = open('E:\\Project\\Saled\\P20190100016-SX1\\R_vs_S\\CLUSTER\\cluster.txt')
# f2 = open('E:\\Project\\Saled\\P20190100016-SX1\\R_vs_S\\CLUSTER\\cluster2.txt', 'w')
# f2.write(f1.readline())
#
# for line in f1:
#     acce = line.split('\t')[0]
#     newline = gene_dict[acce] + '\t' + '\t'.join(line.split('\t')[1:])
#     f2.write(newline)
#
# f1.close()
# f2.close()

### ppi

# gene_df = pd.read_csv('E:\\Project\\Saled\\P20190100016-SX1\\S_vs_C\\gene.txt', sep='\t')
# gene_dict = {k: v for k, v in zip(gene_df['Protein'], gene_df['Gene Name'])}
#
# f1 = open('E:\\Project\\Saled\\P20190100016-SX1\\S_vs_C\\PPI\\p2p_list.txt')
# f2 = open('E:\\Project\\Saled\\P20190100016-SX1\\S_vs_C\\PPI\\p2p_list2.txt', 'w')
# f2.write(f1.readline())
# # print(f1.readlines())
# for line in f1:
#     if re.match(r'\S*\t', line):
#         proteinID = re.match(r'\S*\t', line).group().strip('\t')
#         if proteinID in gene_dict:
#             line.replace(proteinID, str(gene_dict[proteinID]))
#             f2.write(line)
#     else:
#         f2.write(line)
#
# f1.close()
# f2.close()

#####P20181201132-SX1 （添加desription)

# f = open('E:\\Project\\Saled\\P20181201132-SX1\\des.txt')
# des_dict = {}
# for line in f:
#     protein_id = line.split(' ')[0][1:]
#     if re.search(r'description:.*', line):
#         desription = re.search(r'description:.*', line).group().replace('description:', '')
#         des_dict[protein_id] = desription
#     else:
#         des_dict[protein_id] = ''
#
# des_df = pd.DataFrame(index=des_dict.keys(), columns=['Description'])
# des_df['Description'] = des_dict.values()
#
# data1 = pd.read_excel('E:\\Project\\Saled\\P20181201132-SX1\\附件1_蛋白质鉴定列表.xlsx', sheet_name=0, index_col=0)
# data2 = pd.ExcelFile('E:\\Project\\Saled\\P20181201132-SX1\\附件3_蛋白质定量和差异分析列表.xlsx')
# # print(data2.sheet_names)
#
# data1.insert(0, 'Description', des_df.loc[list(data1.index)])       # 指定位置插入列
# data1.to_excel('E:\\Project\\Saled\\P20181201132-SX1\\报告及附件\\附件1_蛋白质鉴定列表.xlsx')
#
# writer = pd.ExcelWriter('E:\\Project\\Saled\\P20181201132-SX1\\报告及附件\\附件3_蛋白质定量和差异分析列表.xlsx')
# for i in data2.sheet_names:
#     df = data2.parse(i, index_col=0)
#     df.insert(0, 'Description', des_df.loc[list(df.index)])
#     df.to_excel(writer, sheet_name=i)
#
# writer.save()
# writer.close()
# f.close()

########P20180700444-SX#########
# group = 'AOT_vs_AST'
# with open('E:\\Project\\Saled\\P20180700444-SX1\\TopBlastHits') as f:
#     f.readline()
#     d = {}
#     for line in f:
#         line = line.strip().split('\t')
#         ID = line[0]
#         if re.search('GN=\S+', line[2]):
#             gene = re.search('GN=\S+', line[2]).group().split('=')[1]
#         else:
#             gene = ID
#         d[ID] = gene

    # data = xlrd.open_workbook(f'E:\\Project\\Saled\\P20180700444-SX2\\{group}\\go\\GO.xlsx')
    # table = data.sheet_by_index(3)
    #
    # wb = Workbook()
    # ws = wb.active
    # ws.append(table.row_values(0))
    # # ws.append(go_table.row_values(1))
    #
    # for row in range(1, table.nrows):
    #     line = table.row_values(row)
    #     test, ref = line[10], line[11]
    #     newtest = ','.join([d.get(i) if i in d else i for i in test.split(',')])
    #     newref = ','.join([d.get(i) if i in d else i for i in ref.split(',')])
    #     line[10], line[11] = newtest, newref
    #     ws.append(line)
    #
    # wb.save(f'E:\\Project\\Saled\\P20180700444-SX2\\{group}\\go\\GO2.xlsx')

    # data2 = xlrd.open_workbook(f'E:\\Project\\Saled\\P20180700444-SX2\\{group}\\kegg\\kegg.xlsx')
    # k_table = data2.sheet_by_index(2)
    #
    # wb = Workbook()
    # ws = wb.active
    # ws.append(k_table.row_values(0))
    # # ws.append(go_table.row_values(1))
    #
    # for row in range(1, k_table.nrows):
    #     line = k_table.row_values(row)
    #     test, ref = line[8], line[9]
    #     newtest = ','.join([d.get(i) if i in d else i for i in test.split(' ')])
    #     newref = ','.join([d.get(i) if i in d else i for i in ref.split(' ')])
    #     line[8], line[9] = newtest, newref
    #     ws.append(line)
    #
    # wb.save(f'E:\\Project\\Saled\\P20180700444-SX2\\{group}\\kegg\\kegg2.xlsx')


    # data = xlrd.open_workbook('E:\\Project\\Saled\\P20180700444-SX1\\质谱鉴定和定量结果\\附件3_蛋白质定量和差异分析列表.xlsx')
    # table = data.sheet_by_index(2)
    #
    # wb = Workbook()
    # ws = wb.active
    # ws.append(table.row_values(0) + ['GeneName'])
    # # ws.append(go_table.row_values(1))
    #
    # for row in range(1, table.nrows):
    #     line = table.row_values(row)
    #     proteinID = line[0]
    #     # # print(proteinID)
    #     geneName = d.get(proteinID)
    #     line.append(geneName)
    #     ws.append(line)
    #
    # wb.save('E:\\Project\\Saled\\P20180700444-SX1\\质谱鉴定和定量结果\\GO2.xlsx')

group = 'AST_vs_AWT'

df = pd.read_excel('E:\\Project\\Saled\\P20180700444-SX2\\补充GO及KEGG表格.xlsx', sheet_name=f'{group}')

go = xlrd.open_workbook(f'E:\\Project\\Saled\\P20180700444-SX2\\{group}\\go\\GO.xlsx')
kegg = xlrd.open_workbook(f'E:\\Project\\Saled\\P20180700444-SX2\\{group}\\kegg\\kegg.xlsx')

def go_func(index):
    go_table = go.sheet_by_index(index)
    go_list = []
    flag1 = 0
    for i in df['Accession']:
        tmp_list1 = []
        for row in range(1, go_table.nrows):
            line = go_table.row_values(row)
            if i in line[5]:
                flag1 = 1
                tmp_list1.append(line[2])
        if flag1 == 0:
            go_list.append('')
        else:
            go_list.append(','.join(tmp_list1))
    return go_list

bp = go_func(0)
mf = go_func(1)
cc = go_func(2)

def kegg_func():
    kegg_table = kegg.sheet_by_index(0)
    ko_list, gene_list, pathway_list = [], [], []
    flag2 = 0
    for i in df['Accession']:
        tmp_list2, tmp_list3, tmp_list4 = [], [], []
        for row in range(1, kegg_table.nrows):
            line2 = kegg_table.row_values(row)
            if i in line2[0]:
                tmp_list2.append(line2[1])
                tmp_list3.append(line2[2])
                tmp_list4.append(line2[5])
                flag2 = 1
        if flag2 == 0:
            ko_list.append('')
            gene_list.append('')
            pathway_list.append('')
        else:
            ko_list.append(','.join(set(tmp_list2)))
            gene_list.append(','.join(set(tmp_list3)))
            pathway_list.append(','.join(set(tmp_list4)))

    return ko_list, gene_list, pathway_list

ko = kegg_func()[0]
gene = kegg_func()[1]
pathway = kegg_func()[2]

# print(len(df['Accession']))
# print(len(ko))

tmp_dict = {'bp': bp,
            'cc': cc,
            'mf': mf,
            'ko': ko,
            'gene': gene,
            'pathway': pathway}
df2 = pd.DataFrame(tmp_dict)
# print(df)
df2.to_excel('E:\\Project\\Saled\\P20180700444-SX2\\diff.xlsx', index=None)


