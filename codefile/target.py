# import sys
import re
import os
import glob
import requests
# import pandas as pd
from collections import defaultdict
from openpyxl import Workbook

def target(path, tmp_type):
    tmp_dict = {'能量代谢': 'n', u'神经递质': 's', u'氨基酸及其衍生物': 'a', u'植物激素': 'z', u'胆汁酸': 'd', 'P200': 'p'}
    os.system('Rscript {} {} {}'.format(os.path.join(os.getcwd(), 'res\\Target-auto.R'), tmp_dict[tmp_type], path))

def groupvsList(path):
    with open(os.path.join(path, 'groupvs.txt')) as fg:
        groupvs_list = [i.strip() for i in fg]
        return groupvs_list

def kegg(path, species):
    with open(os.path.join(os.getcwd(), f'res\\database\\species_specific\\kegg_{species}.Compound')) as fs:
        specific = fs.read()

    with open(os.path.join(os.getcwd(), 'res\\database\\species_specific\\br08901.list')) as fb:
        fb_dict = {line.strip().split('\t')[0]: '\t'.join(line.strip().split('\t')[1:]) for line in fb}

    with open(os.path.join(os.getcwd(), 'res\\database\\metabo.txt')) as fm:
        fm_dict = {}
        for line in fm:
            line = line.strip().split('\t')
            if len(line) > 1:
                fm_dict[line[0]] = line[1]
            else:
                fm_dict[line[0]] = ''

    for groupvs in groupvsList(path):
        fd = open(os.path.join(path, groupvs, 'KEGG', 'diff.txt'))
        fq = open(os.path.join(path, groupvs, 'KEGG', 'query.cpd'), 'w')
        query_cpd_dict = {}
        for line in fd:
            fq.write(f'{line.strip()}\t{fm_dict.get(line.strip())}\n')
            query_cpd_dict[line.strip()] = fm_dict.get(line.strip())

        wb = Workbook()
        ws1 = wb.create_sheet(0, 'query2map')
        ws2 = wb.create_sheet(1, 'map2query')
        ws3 = wb.create_sheet(2, 'IDmapping')

        ws1.append(['Metabolite', 'cpdID', 'cpdName', 'Map_ID', 'Map_Name', 'URL'])
        ws2.append('Pathway_Hierarchy1\tPathway_Hierarchy2\tMap_ID\tMap_Name\tcpdName\tcpdID\tNum_cpd\tURL'.split('\t'))
        ws3.append(['Metabolite', 'C_ID'])
        # with open(os.path.join(path, groupvs, 'KEGG', 'query2map.txt'), 'w') as f1:
        f1 = open(os.path.join(path, groupvs, 'KEGG', 'query2map.txt'), 'w')
        f1.write('Metabolite\tcpdID\tcpdName\tMap_ID\tMap_Name\tURL\n')
            # with open(os.path.join(path, groupvs, 'KEGG', 'map2query.txt'), 'w') as f2:
        f2 = open(os.path.join(path, groupvs, 'KEGG', 'map2query.txt'), 'w')
        f2.write('Pathway_Hierarchy1\tPathway_Hierarchy2\tMap_ID\tMap_Name\tcpdName\tcpdID\tNum_cpd\tURL\n')
        cpdID_list, cpdName_list = defaultdict(list), defaultdict(list)

        for i, j in query_cpd_dict.items():
            if j:
                ws3.append(f'{i}\t{j}'.split('\t'))
                url = f'http://rest.kegg.jp/get/cpd:{j}'
                tmp = requests.get(url).text.replace('\n', '')  #替换全文中的换行符
                ##代谢物名称
                name_com = re.compile('NAME(.*)FORMULA')
                cpdName = name_com.search(tmp).group(1).replace(' ', '')
                # ##代谢物通路
                pathway_com = re.compile('PATHWAY(.*)ENZYME')
                pathway_search = pathway_com.search(tmp)
                if pathway_search:
                    pathwayList = pathway_search.group(1).split('   ')
                    pathwayList = [i.strip() for i in pathwayList if i]
                    for p in pathwayList:
                        p = p.split('  ')
                        map_id = p[0]
                        ko_id = map_id.replace('map', 'ko')
                        if ko_id in specific:
                            map_name = p[1]
                            cpdID_list[map_id].append(j)
                            cpdName_list[map_name].append(cpdName)
                            map_url = f'http://www.kegg.jp/kegg-bin/show_pathway?{map_id}+{j}'
                            f1.write(f'{i}\t{j}\t{cpdName}\t{map_id}\t{map_name}\t{map_url}\n')
                            ws1.append(f'{i}\t{j}\t{cpdName}\t{map_id}\t{map_name}\t{map_url}'.split('\t'))
                else:
                    f1.write(f'{i}\t{j}\t{cpdName}\n')
                    ws1.append(f'{i}\t{j}\t{cpdName}'.split('\t'))
            else:
                f1.write(f'{i}\n')
                ws1.append(f'{i}\t'.split('\t'))
                ws3.append(f'{i}\t'.split('\t'))
        for k, v in zip(cpdID_list.items(), cpdName_list.items()):
            pathway1 = fb_dict.get(k[0]).split('\t')[-1]
            pathway2 = fb_dict.get(k[0]).split('\t')[1]
            new_cpdName = '|'.join(v[1])
            new_cpdID = '+'.join(k[1])
            numcpd = str(len(k[1]))
            url = f'http://www.kegg.jp/kegg-bin/show_pathway?{k[0]}+{new_cpdID}'
            f2.write(f'{pathway1}\t{pathway2}\t{k[0]}\t{v[0]}\t{new_cpdName}\t{new_cpdID}\t{numcpd}\t{url}\n')
            ws2.append(f'{pathway1}\t{pathway2}\t{k[0]}\t{v[0]}\t{new_cpdName}\t{new_cpdID}\t{numcpd}\t{url}'.split('\t'))
        wb.save(os.path.join(path, groupvs, 'KEGG', 'kegg.xlsx'))
        fd.close()
        fq.close()
        f1.close()
        f2.close()

def download_map(path):
    for groupvs in groupvsList(path):
        os.system('Rscript {} {}'.format(os.path.join(os.getcwd(), 'res\\kegg_html_meta.R'), f'{path}\\{groupvs}\\KEGG\\'))

def check(path):
    for groupvs in groupvsList(path):
        with open(os.path.join(path, groupvs, 'KEGG', 'map2query.txt')) as f:
            f.readline()
            c1 = 0
            map_list1 = [i.split('\t')[2] for i in f]
            map_list2 = [os.path.basename(i).split('.')[0] for i in glob.glob(os.path.join(path, groupvs, 'KEGG', 'map', '*.png'))]

            check = set(map_list1) - set(map_list2)
            if check:
                return check
            else:
                return 'OK'

def CpReport(path):
    os.system(f'md {path}\\报告及附件')
    if os.path.exists(f'{path}\\QC_RSD.tiff'):
        os.system(f'copy {path}\\QC_RSD.tiff {path}\\报告及附件')

    for groupvs in groupvsList(path):
        os.system(f' md {path}\\报告及附件\\{groupvs}\\Metaboanalyst_Data')
        # os.system(f'md {path}\\报告及附件\\{groupvs}\\Metaboanalyst_Data\\Heatmap')
        # os.system(f'md {path}\\报告及附件\\{groupvs}\\Metaboanalyst_Data\\Overview')
        os.system(f'copy {path}\\{groupvs}\\data_metaboanalyst.xls {path}\\报告及附件\\{groupvs}')
        os.system(f'copy {path}\\{groupvs}\\flitered_data_metaboanalyst.xls {path}\\报告及附件\\{groupvs}')
        os.system(f'xcopy {path}\\{groupvs}\\Metaboanalyst_Data\\Heatmap {path}\\报告及附件\\{groupvs}\\Metaboanalyst_Data\\Heatmap\\')
        os.system(f'xcopy {path}\\{groupvs}\\Metaboanalyst_Data\\Overview {path}\\报告及附件\\{groupvs}\\Metaboanalyst_Data\\Overview\\')
        os.system(f'xcopy {path}\\{groupvs}\\Metaboanalyst_Data\\Boxplot {path}\\报告及附件\\{groupvs}\\Metaboanalyst_Data\\Boxplot\\')

        if os.path.exists(f'{path}\\{groupvs}\\KEGG'):
            os.system(f'md {path}\\报告及附件\\{groupvs}\\KEGG')
            os.system(f'copy {path}\\{groupvs}\\KEGG\\kegg.xlsx {path}\\报告及附件\\{groupvs}\\KEGG\\')
            os.system(f'xcopy {path}\\{groupvs}\\KEGG\\map {path}\\报告及附件\\{groupvs}\\KEGG\\map\\')
        

