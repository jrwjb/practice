
from openpyxl import Workbook

# with open('E:\\Project\\Saled\\P15610-SX3\\GO.annot') as f:
#     with open('E:\\Project\\Saled\\P15610-SX3\\PPI\\module\\Cluster_1.txt') as f1:
#         with open('E:\\Project\\Saled\\P15610-SX3\\PPI\\module\\Cluster_2.txt') as f2:
#             with open('E:\\Project\\Saled\\P15610-SX3\\PPI\\module\\Cluster_3.txt') as f3:
#                 with open('E:\\Project\\Saled\\P15610-SX3\\PPI\\module\\Cluster_3.txt') as f4:
#                     with open('E:\\Project\\Saled\\P15610-SX3\\PPI\\module\\Cluster_3.txt') as f5:
#         # with open('E:\\Project\\Saled\\P20181000883-SX2\\12H_vs_0H\\PPI\\module\\cluster.csv', 'w') as fw:
#                         mylist1 = [i.strip() for i in f1]
#                         mylist2 = [i.strip() for i in f2]
#                         mylist3 = [i.strip() for i in f3]
#                         mylist4 = [i.strip() for i in f4]
#                         mylist5 = [i.strip() for i in f5]
#
#                         wb = Workbook()
#                         ws1 = wb.create_sheet('Cluster1_list', index=0)
#                         ws2 = wb.create_sheet('Cluster2_list', index=1)
#                         ws3 = wb.create_sheet('Cluster3_list', index=2)
#                         ws4 = wb.create_sheet('Cluster4_list', index=3)
#                         ws5 = wb.create_sheet('Cluster5_list', index=4)
#                         title = ['Protein ID', 'GO annotations']
#                         ws1.append(title)
#                         ws2.append(title)
#                         ws3.append(title)
#                         ws4.append(title)
#                         ws5.append(title)
#                         # print(f.readline())
#                         # print(f.readline().strip().split('\t'))
#                         # fw.write('\t'.join(f.readline().strip().split('\t')) + '\n')
#                         for i in f:
#                             i = i.strip()
#                             id = '.'.join(i.split('\t')[0].split('.')[0:2])
#                             go = i.split('\t')[1] + '(' + i.split('\t')[3] + ':' + i.split('\t')[2] + ')'
#                             # print(go)
#                             # id = i.strip().split()[0]
#
#                             if id in mylist1:
#                                 s = id + '\t' + go
#                                 ws1.append(s.split('\t'))
#
#                             elif id in mylist2:
#                                 s = id + '\t' + go
#                                 ws2.append(s.split('\t'))
#                             elif id in mylist3:
#                                 s = id + '\t' + go
#                                 ws3.append(s.split('\t'))
#                             elif id in mylist4:
#                                 s = id + '\t' + go
#                                 ws4.append(s.split('\t'))
#                             elif id in mylist5:
#                                 s = id + '\t' + go
#                                 ws5.append(s.split('\t'))
#                         wb.save('E:\\Project\\Saled\\P15610-SX3\\PPI\\module\\Clusters.xlsx')


import pandas as pd

pd.set_option('display.max_columns', None)

df = pd.read_csv('E:\\Project\\Saled\\P15610-SX3\\GO.annot', header=None, sep='\t')
f1 = open('E:\\Project\\Saled\\P15610-SX3\\PPI\\module\\Cluster_1.txt')
f2 = open('E:\\Project\\Saled\\P15610-SX3\\PPI\\module\\Cluster_2.txt')
f3 = open('E:\\Project\\Saled\\P15610-SX3\\PPI\\module\\Cluster_3.txt')
f4 = open('E:\\Project\\Saled\\P15610-SX3\\PPI\\module\\Cluster_4.txt')
f5 = open('E:\\Project\\Saled\\P15610-SX3\\PPI\\module\\Cluster_5.txt')

mylist1 = [i.strip() for i in f1]
mylist2 = [i.strip() for i in f2]
mylist3 = [i.strip() for i in f3]
mylist4 = [i.strip() for i in f4]
mylist5 = [i.strip() for i in f5]

wb = Workbook()
ws1 = wb.create_sheet('Cluster1_list', index=0)
ws2 = wb.create_sheet('Cluster2_list', index=1)
ws3 = wb.create_sheet('Cluster3_list', index=2)
ws4 = wb.create_sheet('Cluster4_list', index=3)
ws5 = wb.create_sheet('Cluster5_list', index=4)
title = ['Protein ID', 'GO annotations']
ws1.append(title)
ws2.append(title)
ws3.append(title)
ws4.append(title)
ws5.append(title)

df_group = df.groupby(0)
for i in df_group:
    id = i[0]
    df_tmp = i[1]
    mylist = []
    for j in range(df_tmp.shape[0]):
        go = str(df_tmp.iloc[j, 1]) + '(' + str(df_tmp.iloc[j, 3]) + ':' + str(df_tmp.iloc[j, 2]) + ')'
        mylist.append(go)
    # print(id)
    # print(mylist)
    tmp_list = []
    if id in mylist1:
        tmp_list.append(id)
        tmp_list.append(';'.join(mylist))
        ws1.append(tmp_list)
    elif id in mylist2:
        tmp_list.append(id)
        tmp_list.append(';'.join(mylist))
        ws2.append(tmp_list)
    elif id in mylist3:
        tmp_list.append(id)
        tmp_list.append(';'.join(mylist))
        ws3.append(tmp_list)
    elif id in mylist4:
        tmp_list.append(id)
        tmp_list.append(';'.join(mylist))
        ws4.append(tmp_list)
    elif id in mylist5:
        tmp_list.append(id)
        tmp_list.append(';'.join(mylist))
        ws5.append(tmp_list)
wb.save('E:\\Project\\Saled\\P15610-SX3\\PPI\\module\\Clusters.xlsx')

f1.close()
f2.close()
f3.close()
f4.close()
f5.close()