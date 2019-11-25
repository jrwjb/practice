import sys
import re
import os
import requests
import pandas as pd
from collections import defaultdict
from openpyxl import Workbook

# path = sys.argv[1]
path = 'C:\\Users\\jbwang\\Desktop\\tq\\target\\demo'
with open(os.path.join(path, 'folders.txt')) as fg:
    groupvs_list = [i.strip() for i in fg]

with open('C:\\Users\\jbwang\\Desktop\\tq\\database\\species_specific\\kegg_ko.Compound') as fs:
    specific = fs.read()

with open('C:\\Users\\jbwang\\Desktop\\tq\\database\\species_specific\\br08901.list') as fb:
    fb_dict = {line.strip().split('\t')[0]: '\t'.join(line.strip().split('\t')[1:]) for line in fb}

metabo_df = pd.read_csv('C:\\Users\\jbwang\\Desktop\\tq\\database\\metabo.txt', sep='\t')

def query2map(query_cpd):
    # url = 'https://www.kegg.jp/dbget-bin/www_bget?cpd:C12270'
    url = 'https://www.kegg.jp/dbget-bin/www_bget?cpd:{}'.format(query_cpd)
    tmp = requests.get(url).text.replace('\n', '')  #替换全文中的换行符
    ##代谢物名称
    name_com = re.compile('<nobr>Name</nobr>.+hidden">(.+?)<br></div></div></td></tr>')
    name = name_com.search(tmp).group(1).replace('<br>', '')
    ##代谢物通路
    pathway_com = re.compile('<a href="/kegg-bin/show_pathway\?(\w+)\+(\w+)".+?<td>(.+?)</td>')
    pathway_list = pathway_com.findall(tmp)
    # print(pathway_list)
    return name, pathway_list

def kegg(groupvs):
    diff_df = pd.read_csv(os.path.join(path, groupvs, 'KEGG', 'diff.txt'), sep='\t', header=None)
    diff_df.columns = ['name']
    query_cpd = pd.merge(diff_df, metabo_df, how='left', on='name')
    query_cpd.to_csv(os.path.join(path, groupvs, 'KEGG', 'query.cpd'), index=False, header=None, sep='\t')

    wb = Workbook()
    ws1 = wb.create_sheet('query2map', 0)
    ws2 = wb.create_sheet('map2query', 1)
    ws3 = wb.create_sheet('IDmapping', 2)

    ws1.append(['Metabolite', 'cpdID', 'cpdName', 'Map_ID', 'Map_Name', 'URL'])
    ws2.append('Pathway_Hierarchy1\tPathway_Hierarchy2\tMap_ID\tMap_Name\tcpdName\tcpdID\tNum_cpd\tURL'.split('\t'))
    ws3.append(['Metabolite', 'C_ID'])
    with open(os.path.join(path, groupvs, 'KEGG', 'query2map.txt'), 'w') as f1:
        f1.write('Metabolite\tcpdID\tcpdName\tMap_ID\tMap_Name\tURL\n')
        with open(os.path.join(path, groupvs, 'KEGG', 'map2query.txt'), 'w') as f2:
            f2.write('Pathway_Hierarchy1\tPathway_Hierarchy2\tMap_ID\tMap_Name\tcpdName\tcpdID\tNum_cpd\tURL\n')
            cpdID_list, cpdName_list = defaultdict(list), defaultdict(list)
            for i, j in zip(query_cpd['name'], query_cpd['KEGG ID']):
                if str(j) == 'nan':
                    f1.write(f'{i}\n')
                    ws1.append(f'{i}\t'.split('\t'))
                    ws3.append(f'{i}\t'.split('\t'))
                else:
                    ws3.append(f'{i}\t{j}'.split('\t'))
                    cpdName = query2map(j)[0]
                    pathwayList = query2map(j)[1]
                    if len(pathwayList) == 0:
                        f1.write(f'{i}\t{j}\t{cpdName}\n')
                        ws1.append(f'{i}\t{j}\t{cpdName}'.split('\t'))
                    else:
                        for p in pathwayList:
                            map_id = p[0]
                            ko_id = map_id.replace('map', 'ko')
                            if ko_id in specific:
                                cpdID = p[1]
                                map_name = p[2]
                                cpdID_list[map_id].append(cpdID)
                                cpdName_list[map_name].append(cpdName)
                                map_url = f'http://www.kegg.jp/kegg-bin/show_pathway?{map_id}+{j}'
                                f1.write(f'{i}\t{j}\t{cpdName}\t{map_id}\t{map_name}\t{map_url}\n')
                                ws1.append(f'{i}\t{j}\t{cpdName}\t{map_id}\t{map_name}\t{map_url}'.split('\t'))
            for k, v in zip(cpdID_list.items(), cpdName_list.items()):
                pathway1 = fb_dict.get(k[0]).split('\t')[-1]
                pathway2 = fb_dict.get(k[0]).split('\t')[1]
                new_cpdName = '|'.join(v[1])
                new_cpdID = '+'.join(k[1])
                numcpd = str(len(k[1]))
                url = f'http://www.kegg.jp/kegg-bin/show_pathway?{k[0]}+{new_cpdID}'
                f2.write(f'{pathway1}\t{pathway2}\t{k[0]}\t{v[0]}\t{new_cpdName}\t{new_cpdID}\t{numcpd}\t{url}\n')
                ws2.append(f'{pathway1}\t{pathway2}\t{k[0]}\t{v[0]}\t{new_cpdName}\t{new_cpdID}\t{numcpd}\t{url}'.split('\t'))
    wb.save(os.path.join(path, groupvs, 'KEGG', 'kegg.xlsx'))

def main():
    for g in groupvs_list:
        kegg(g)


if __name__ == '__main__':
    main()
